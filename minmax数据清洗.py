
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

 

# 每个指标的配置：(文件名, Sheet名, 表头行, 指标中文名, 正向/负向)
#   header_row : pandas read_excel 的 header 参数（0=第一行，1=第二行，None=无表头）
#   direction  : "positive"（正向，越大越好）或 "negative"（负向，越小越好）

INDICATORS = [
    {
        "file":       r"D:\用户\Desktop\指标全部\001基本养老保险覆盖率及参保人数.xlsx",
        "sheet":      "基本养老保险覆盖率",
        "header_row": 1,           
        "name":       "基本养老保险覆盖率",
        "direction":  "positive",
    },
    {
        "file":       r"D:\用户\Desktop\指标全部\002基本养老保险基金收入强度.xlsx",
        "sheet":      "基本养老保险基金收入强度",
        "header_row": 0,
        "name":       "基本养老保险基金收入强度",
        "direction":  "positive",
    },
    {
        "file":       r"D:\用户\Desktop\指标全部\003基本养老保险基金可支付月数.xlsx",
        "sheet":      "基本养老保险基金可支付月数",
        "header_row": 1,           
        "name":       "基本养老保险基金可支付月数",
        "direction":  "positive",
    },
    {
        "file":       r"D:\用户\Desktop\指标全部\004城镇职工养老金水平.xlsx",
        "sheet":      "城镇职工养老金水平",
        "header_row": 1,
        "name":       "城镇职工养老金水平",
        "direction":  "positive",
    },
    {
        "file":       r"D:\用户\Desktop\指标全部\005城乡居民养老金水平.xlsx",
        "sheet":      "城乡居民养老金水平",
        "header_row": 1,
        "name":       "城乡居民养老金水平",
        "direction":  "positive",
    },
    {
        "file":       r"D:\用户\Desktop\指标全部\006企业年金覆盖率.xlsx",
        "sheet":      "企业年金覆盖率",
        "header_row": 1,
        "name":       "企业年金覆盖率",
        "direction":  "positive",
    },
    {
        "file":       r"D:\用户\Desktop\指标全部\007企业年金基金积累强度.xlsx",
        "sheet":      "企业年金基金积累强度",
        "header_row": 1,
        "name":       "企业年金基金积累强度",
        "direction":  "positive",
    },
    {
        "file":       r"D:\用户\Desktop\指标全部\008健康保险密度.xlsx",
        "sheet":      "健康保险密度",
        "header_row": 1,
        "name":       "健康保险密度",
        "direction":  "positive",
    },
    {
        "file":       r"D:\用户\Desktop\指标全部\009人寿保险密度.xlsx",
        "sheet":      "人寿保险密度",
        "header_row": 1,
        "name":       "人寿保险密度",
        "direction":  "positive",
    },
    {
        "file":       r"D:\用户\Desktop\指标全部\011index_aggregate_wide.xlsx",
        "sheet":      "Sheet1",
        "header_row": 0,
        "name":       "省级数字普惠金融指数",
        "direction":  "positive",
    },
    {
        "file":       r"D:\用户\Desktop\指标全部\012全国分省每万人银行网点数.xlsx",
        "sheet":      "Sheet1",
        "header_row": 0,
        "name":       "银行业金融机构网点密度（每万人）",
        "direction":  "positive",
    },
    {
        "file":       r"D:\用户\Desktop\指标全部\013养老PPP投资强度0416.xlsx",
        "sheet":      "Sheet1",
        "header_row": 0,
        "name":       "养老服务类PPP投资强度",
        "direction":  "positive",
        
    },
    {
        "file":       r"D:\用户\Desktop\指标全部\014卫生和社会工作固定资产投资增速.xlsx",
        "sheet":      "Sheet1",
        "header_row": 0,
        "name":       "卫生和社会工作固定资产投资增速",
        "direction":  "positive",
        
    },
    
    
    

  
]



OUTPUT_FILE = "养老金融高质量发展省级面板指标_MinMax标准化.xlsx"

# ============================================================

# ============================================================

def load_indicator(cfg: dict) -> pd.DataFrame:
    """
    读取单个指标文件，统一返回格式：
        第一列 = '地区'，其余列 = 年份字符串（如 '2024'）
    """
    df = pd.read_excel(cfg["file"], sheet_name=cfg["sheet"],
                       header=cfg["header_row"])

    
    new_cols = []
    for col in df.columns:
        s = str(col).strip().replace("年", "")
        new_cols.append(s)
    df.columns = new_cols

    # 第一列统一命名为"地区"
    df.rename(columns={df.columns[0]: "地区"}, inplace=True)

    # 年份列转数值
    year_cols = [c for c in df.columns if c != "地区"]
    df[year_cols] = df[year_cols].apply(pd.to_numeric, errors="coerce")

    # 删除全为空的行
    df.dropna(subset=year_cols, how="all", inplace=True)
    df.reset_index(drop=True, inplace=True)

    return df


def minmax_normalize(df: pd.DataFrame, direction: str) -> tuple:
    """
    对 df（地区×年份）做全样本极差标准化。
    正向：(X - min) / (max - min)
    负向：(max - X) / (max - min)
    返回 (标准化后的df, global_min, global_max)
    """
    year_cols = [c for c in df.columns if c != "地区"]
    vals = df[year_cols].values.astype(float)

    gmin = np.nanmin(vals)
    gmax = np.nanmax(vals)
    rng  = gmax - gmin

    df_norm = df.copy()
    if direction == "positive":
        df_norm[year_cols] = (vals - gmin) / rng
    else:  # negative
        df_norm[year_cols] = (gmax - vals) / rng

    df_norm[year_cols] = df_norm[year_cols].round(8)
    return df_norm, gmin, gmax


# ============================================================

# ============================================================

THEME_COLORS = [
    ("1F4E79", "D6E4F0"),  # 蓝
    ("375623", "D5E8D4"),  # 绿
    ("843C0C", "FFE6CC"),  # 橙
    ("4A235A", "E8D5F5"),  # 紫
    ("1B4F72", "D1F2EB"),  # 青
    ("6E2C00", "FDEBD0"),  # 棕
    ("145A32", "D5F5E3"),  # 深绿
    ("4D4D4D", "F2F2F2"),  # 灰
]


def write_indicator_sheet(wb, cfg, df_orig, df_norm, gmin, gmax, color_idx):
    header_hex, row_hex = THEME_COLORS[color_idx % len(THEME_COLORS)]
    year_cols = [c for c in df_orig.columns if c != "地区"]
    n_cols    = len(year_cols)
    n_rows    = len(df_orig)
    last_col  = get_column_letter(n_cols + 1)
    is_ratio  = (gmax <= 2)  # 判断是比率还是绝对值

    sname = cfg["name"][:28]  # sheet名最长31字符
    ws = wb.create_sheet(sname)

    hfill  = PatternFill("solid", start_color=header_hex)
    rfill  = PatternFill("solid", start_color=row_hex)
    w_fill = PatternFill("solid", start_color="FFFFFF")
    z_fill = PatternFill("solid", start_color="F9F9F9")
    g_fill = PatternFill("solid", start_color="F2F2F2")

    hfont  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    sfont  = Font(name="Arial", bold=True, size=9)
    cfont  = Font(name="Arial", size=9)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    def set_header(cell, val):
        cell.value, cell.font, cell.fill, cell.alignment = val, hfont, hfill, center

    def set_subheader(cell, val):
        cell.value, cell.font, cell.fill, cell.alignment = val, sfont, g_fill, center

    # ── 标题行
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    dir_str = "正向↑" if cfg["direction"] == "positive" else "负向↓"
    c.value = f"{cfg['name']}  |  Min-Max 极差标准化  |  {dir_str}  |  全样本固定端点"
    c.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    c.fill  = hfill
    c.alignment = center
    ws.row_dimensions[1].height = 22

    # ── 信息行
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    fmt = ".6f" if is_ratio else ",.0f"
    c.value = (f"全样本最小值 = {gmin:{fmt}}    "
               f"全样本最大值 = {gmax:{fmt}}    ")
    c.font  = Font(name="Arial", italic=True, size=9, color="595959")
    c.alignment = left

    # ── 原始值区块
    def write_block(start_row, df_block, title, num_fmt):
        # 区块标题
        ws.merge_cells(f"A{start_row}:{last_col}{start_row}")
        set_header(ws[f"A{start_row}"], title)
        ws.row_dimensions[start_row].height = 18

        # 列头
        r = start_row + 1
        set_subheader(ws.cell(r, 1), "地区")
        for j, yr in enumerate(year_cols):
            set_subheader(ws.cell(r, j + 2), yr)

        # 数据
        for i, row in df_block.iterrows():
            r = start_row + 2 + i
            fill = w_fill if i % 2 == 0 else z_fill
            c = ws.cell(r, 1, row["地区"])
            c.font, c.fill, c.alignment = cfont, fill, left
            for j, yr in enumerate(year_cols):
                cell = ws.cell(r, j + 2, row[yr])
                cell.font, cell.fill, cell.alignment = cfont, fill, center
                cell.number_format = num_fmt

    raw_fmt  = "0.000000" if is_ratio else "#,##0"
    norm_fmt = "0.0000"

    write_block(3,                    df_orig, "▌ 原始值",               raw_fmt)
    write_block(3 + n_rows + 4,       df_norm, "▌ Min-Max 标准化值 [0,1]", norm_fmt)

    # ── 列宽
    ws.column_dimensions["A"].width = 16
    for j in range(n_cols):
        ws.column_dimensions[get_column_letter(j + 2)].width = 11
    ws.freeze_panes = "B5"


def write_summary_sheet(wb, results):
    ws = wb.create_sheet("汇总说明", 0)
    hfill = PatternFill("solid", start_color="1F4E79")

    ws["A1"] = "银发经济指标体系 — Min-Max 极差标准化汇总"
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    ws.merge_cells("A1:G1")

    headers = ["指标名称", "文件", "方向", "全样本min", "全样本max", "年份范围", "省份数"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(3, j, h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = hfill
        c.alignment = Alignment(horizontal="center")

    for i, (cfg, df_orig, gmin, gmax) in enumerate(results, 4):
        year_cols = [c for c in df_orig.columns if c != "地区"]
        is_ratio  = gmax <= 2
        fmt = f"{gmin:.6f}" if is_ratio else f"{gmin:,.0f}"
        fmx = f"{gmax:.6f}" if is_ratio else f"{gmax:,.0f}"
        row = [cfg["name"], cfg["file"],
               "正向↑" if cfg["direction"] == "positive" else "负向↓",
               fmt, fmx,
               f"{min(year_cols)}–{max(year_cols)}",
               len(df_orig)]
        fill = PatternFill("solid", start_color="DEEAF1" if i % 2 == 0 else "FFFFFF")
        for j, val in enumerate(row, 1):
            c = ws.cell(i, j, val)
            c.font = Font(name="Arial", size=9)
            c.fill = fill

    # 方法说明
    notes = [
        (i + len(results) + 5, t) for i, t in enumerate([
            "【标准化方法说明】",
            "方法：Min-Max 极差标准化",
            "正向公式：标准化值 = (X - X_min) / (X_max - X_min)",
            "负向公式：标准化值 = (X_max - X) / (X_max - X_min)",
            "端点取值：各指标全样本（所有省份 × 所有年份）固定min/max，保证跨年纵向可比",
            "输出区间：[0, 1]，0 = 全样本最低，1 = 全样本最高",
        ])
    ]
    for row_num, text in notes:
        c = ws.cell(row_num, 1, text)
        c.font = Font(name="Arial", bold=(text.startswith("【")), size=9)
        ws.merge_cells(f"A{row_num}:G{row_num}")

    for col, width in zip("ABCDEFG", [28, 22, 8, 14, 14, 16, 8]):
        ws.column_dimensions[col].width = width


# ============================================================
#  主程序
# ============================================================

def main():
    print("=" * 55)
    print("  Min-Max 极差标准化  —  养老金融高质量发展省级面板指标")
    print("=" * 55)

    wb = Workbook()
    wb.remove(wb.active)

    results = []
    for idx, cfg in enumerate(INDICATORS):
        print(f"\n[{idx+1}/{len(INDICATORS)}] 处理: {cfg['name']}  ({cfg['file']})")

        df_orig          = load_indicator(cfg)
        df_norm, gmin, gmax = minmax_normalize(df_orig, cfg["direction"])

        year_cols = [c for c in df_orig.columns if c != "地区"]
        is_ratio  = gmax <= 2
        fmt = f"{gmin:.6f}" if is_ratio else f"{gmin:,.0f}"
        fmx = f"{gmax:.6f}" if is_ratio else f"{gmax:,.0f}"
        print(f"    省份: {len(df_orig)}  年份: {year_cols}")
        print(f"    全样本 min={fmt}  max={fmx}")

        write_indicator_sheet(wb, cfg, df_orig, df_norm, gmin, gmax, idx)
        results.append((cfg, df_orig, gmin, gmax))

    write_summary_sheet(wb, results)
    wb.save(OUTPUT_FILE)
    print(f"\n✅ 已保存 → {OUTPUT_FILE}")
    print("=" * 55)


if __name__ == "__main__":
    main()
